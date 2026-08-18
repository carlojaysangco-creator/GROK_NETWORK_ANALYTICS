"""Load topology and mapping tables from local CSV/XLSX files."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from network_analytics.data_platform import GenerationStore, SourceIdentity
from network_analytics.shared.numbers import optional_float

from .ftth import publish_ftth_mapping
from .reference import DATASET_TOPOLOGY, publish_link_topology


def _sha256_file(path: Path) -> str:
    import hashlib

    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_tabular_rows(path: Path, *, sheet: str | None = None) -> list[dict[str, Any]]:
    path = path.expanduser().resolve()
    if not path.is_file():
        raise FileNotFoundError(path)
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt"}:
        text = path.read_text(encoding="utf-8-sig")
        return list(csv.DictReader(text.splitlines()))
    if suffix in {".xlsx", ".xlsm"}:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            if sheet and sheet in wb.sheetnames:
                ws = wb[sheet]
            else:
                preferred = ("weekly_latest_week", "Sheet1")
                chosen = None
                lower = {s.lower(): s for s in wb.sheetnames}
                for name in preferred:
                    if name.lower() in lower:
                        chosen = lower[name.lower()]
                        break
                ws = wb[chosen or wb.sheetnames[0]]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                return []
            keys = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header)]
            out: list[dict[str, Any]] = []
            for row in rows_iter:
                item = {keys[i]: row[i] if i < len(row) else None for i in range(len(keys))}
                if any(v is not None and str(v).strip() != "" for v in item.values()):
                    out.append(item)
            return out
        finally:
            wb.close()
    raise ValueError(f"unsupported tabular format: {suffix}")


def publish_topology_file(
    store: GenerationStore,
    path: Path,
    *,
    producer_version: str,
    sheet: str | None = None,
    promote: bool = True,
) -> object:
    rows = load_tabular_rows(path, sheet=sheet)
    source = SourceIdentity(
        system="file",
        path_or_job=str(path),
        sha256=_sha256_file(path),
    )
    return publish_link_topology(
        store,
        rows,
        producer_version=producer_version,
        source=source,
        promote=promote,
    )


def publish_ftth_file(
    store: GenerationStore,
    path: Path,
    *,
    producer_version: str,
    sheet: str | None = None,
    promote: bool = True,
) -> object:
    rows = load_tabular_rows(path, sheet=sheet)
    source = SourceIdentity(system="file", path_or_job=str(path), sha256=_sha256_file(path))
    return publish_ftth_mapping(
        store,
        rows,
        producer_version=producer_version,
        source=source,
        promote=promote,
    )


def publish_daily_topology_from_fact(
    store: GenerationStore,
    *,
    producer_version: str,
    promote: bool = True,
) -> object:
    """Build rpa_daily_topology from promoted FACT parent/physical rows.

    NetLynx monitoring is not a routing substitute unless explicitly published
    as this daily topology generation.
    """

    from network_analytics.netlynx import load_observations
    from network_analytics.netlynx.contracts import InterfaceType
    from .daily import DATASET_DAILY_TOPOLOGY
    from .reference import publish_link_topology

    observations = load_observations(store)
    rows: list[dict[str, Any]] = []
    for obs in observations:
        if obs.interface_type not in {InterfaceType.LAG_PARENT, InterfaceType.PHYSICAL}:
            continue
        if not obs.a_end or not obs.z_end:
            continue
        util = obs.max_util_pct()
        rows.append(
            {
                "a_end": obs.a_end,
                "z_end": obs.z_end,
                "capacity_mbps": obs.capacity_mbps,
                "max_util_pct": util,
                "in_util_pct": obs.in_util_pct,
                "out_util_pct": obs.out_util_pct,
                "link_id": obs.link_id,
                "role": "parent" if obs.interface_type is InterfaceType.LAG_PARENT else "physical",
                "weight": 1.0,
            }
        )
    if not rows:
        raise ValueError("no authoritative FACT rows with endpoints to build daily topology")

    # Publish into daily dataset name by temporarily using low-level store path
    from network_analytics.data_platform import SourceIdentity, ValidationSummary
    from .link_schema import link_records_from_rows
    import json

    accepted, rejected = link_records_from_rows(rows)
    ref = store.create_candidate(
        dataset_name=DATASET_DAILY_TOPOLOGY,
        schema_version="topology-v1",
        producer_version=producer_version,
        source=SourceIdentity(
            system="derived",
            path_or_job="netlynx_fact→rpa_daily_topology",
            sha256="derived",
        ),
        metadata={"source_dataset": "netlynx_fact", "row_count": len(rows)},
    )
    payload = "\n".join(
        json.dumps(
            {
                "a_end": r.a_end,
                "z_end": r.z_end,
                "weight": r.weight,
                "capacity_mbps": r.capacity_mbps,
                "max_util_pct": r.max_util_pct,
                "link_type": r.link_type,
                "role": r.role.value,
                "link_id": r.link_id,
                "parent_link_id": r.parent_link_id,
                "member_count": r.member_count,
            },
            sort_keys=True,
        )
        for r in accepted
    ).encode("utf-8")
    store.add_data_file(ref, "links.jsonl", payload + (b"\n" if payload else b""))
    store.mark_validated(
        ref,
        input_count=len(rows),
        accepted_count=len(accepted),
        rejected_count=rejected,
        validation=ValidationSummary(),
    )
    store.publish(ref)
    if promote:
        store.promote(DATASET_DAILY_TOPOLOGY, ref.generation_id)
    return store._load_ref(DATASET_DAILY_TOPOLOGY, ref.generation_id)
